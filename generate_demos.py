
from openpyxl import Workbook
import datetime
import random

# Common Data
usernames = [f"user_{i}" for i in range(1, 21)]
headers = ['Username', 'Password', 'LastBackup', 'BackupFrequency', 'BackupType', 'BackupStatus', 'RetentionDays']

def create_sheet(wb, scenario):
    ws = wb.active
    ws.append(headers)
    
    for user in usernames:
        row = [user]
        
        # 1. Password Logic
        if scenario == 'high':
            # Strong passwords
            row.append(f"P@ssw0rd{random.randint(1000,9999)}!Goal")
        elif scenario == 'low':
            # Weak passwords
            row.append("pass")
        else:
            # Mixed
            if random.random() > 0.5:
                 row.append(f"P@ssw0rd{random.randint(1000,9999)}!Goal")
            else:
                 row.append("123456")

        # 2. Backup Logic
        today = datetime.date.today()
        
        if scenario == 'high':
            # Good Backups
            row.append((today - datetime.timedelta(days=random.randint(0, 2))).isoformat())
            row.append('Daily')
            row.append('Full')
            row.append('Success')
            row.append(45)
        elif scenario == 'low':
            # Bad Backups
            row.append((today - datetime.timedelta(days=random.randint(10, 50))).isoformat())
            row.append('Monthly')
            row.append('Differential')
            row.append('Failed')
            row.append(7)
        else:
            # Mixed
            is_good = random.random() > 0.5
            if is_good:
                row.append((today - datetime.timedelta(days=1)).isoformat())
                row.append('Daily')
                row.append('Full')
                row.append('Success')
                row.append(30)
            else:
                row.append((today - datetime.timedelta(days=20)).isoformat())
                row.append('Irregular')
                row.append('Full')
                row.append('Failed')
                row.append(30)
            
        ws.append(row)

# Generate Files
print("Generating 'demo_compliance_high.xlsx'...", flush=True)
wb = Workbook()
create_sheet(wb, 'high')
wb.save('demo_compliance_high.xlsx')
print("Saved 'demo_compliance_high.xlsx'", flush=True)

print("Generating 'demo_compliance_low.xlsx'...", flush=True)
wb = Workbook()
create_sheet(wb, 'low')
wb.save('demo_compliance_low.xlsx')
print("Saved 'demo_compliance_low.xlsx'", flush=True)

print("Generating 'demo_compliance_mixed.xlsx'...", flush=True)
wb = Workbook()
create_sheet(wb, 'mixed')
wb.save('demo_compliance_mixed.xlsx')
print("Saved 'demo_compliance_mixed.xlsx'", flush=True)

print("Done! Three files created.", flush=True)
